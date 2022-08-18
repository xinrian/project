from openpyxl import *
from openpyxl.chart import PieChart, Reference

#  对象层创建一个实例化对象
wb = Workbook()
#  激活表格然后表格操作
test = wb.active
# ws['A1'] = '姓名'
# ws['B1'] = '年龄'
# ws['C1'] = '身高'
# ws['D1'] = '体重'
#
# ws.append(['测试人员', 25, 185, 70])

# sheet_one = wb.create_sheet("少林寺")#创建一个 sheet 名为 sheet
# huoqu_sheetname = wb.get_sheet_names=('少林寺')
#
# sheet_one_copy = wb.copy_worksheet(sheet_one)
#
# print(huoqu_sheetname)

#  按照坐标填写内容
# test = wb.create_sheet('测试数据')
# d = test.cell(3,3,3)

#  按照单个的单元格进行填写数据
# test['A1'] = "测试人员"
# test['B1'] = 22

# test = wb.create_sheet("pycharm")
# for a in range(25):
#     for b in range(25):
#         for c in range(25):
#             test.append((a, b, c))

data = [
    ['水果', '销量'],
    ['苹果', 50],
    ['樱桃', 30],
    ['橘子', 10],
    ['香蕉', 40],
]

wb = Workbook()
ws = wb.active

for row in data:
    ws.append(row)

pie = PieChart()
pie.title = "水果销量占比"
labels = Reference(ws, 1, 2, 5)
data = Reference(ws, 2, 1, 5)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)

ws.add_chart(pie, "D1")

wb.save('excel/excel.xlsx')
#  获取指定文件的sheet页名称
# wb = load_workbook('excel/excel.xlsx')
# print(wb.sheetnames)
